#!/usr/bin/env python3
"""
Parse the "Suicide" workbook — Mike's survivor pool, mailed weekly with every
entrant's pick history.

Layout (Sheet1):

    C1              "Suicide Pool"
    row 2, D..T     week numbers 1..17
    row 3+, A       entry number
    row 3+, B       real name (often blank — nickname is the reliable id)
    row 3+, C       nickname
    row 3+, D..T    team abbreviation picked that week, blank once eliminated
    col V           reference list of team abbreviations (ignored)

Two things the pickem parser doesn't have to deal with:

  * A blank cell is ambiguous. Mid-season it means "eliminated"; for a week
    that hasn't happened yet it means "not submitted". We resolve this by
    treating the last week with ANY picks as the current week.
  * Team abbreviations are hand-typed and inconsistent ("LA R", "LAR").
    ALIASES normalises them; unknown codes are reported rather than silently
    passed through, because a typo here corrupts the used-teams list that the
    whole future-value plan depends on.

    python scripts/parse_survivor.py "path/to/Suicide 25.xlsx" 2025
"""

import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FIRST_WEEK_COL = 4          # column D
MAX_WEEK_COL = 20           # column T = week 17

TEAMS = {
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE", "DAL", "DEN",
    "DET", "GB", "HOU", "IND", "JAC", "KC", "LAC", "LAR", "LV", "MIA",
    "MIN", "NE", "NO", "NYG", "NYJ", "PHI", "PIT", "SEA", "SF", "TB",
    "TEN", "WAS",
}

# Hand-typed variants seen in the workbook, plus the obvious near-misses.
ALIASES = {
    "LA R": "LAR", "LAR": "LAR", "RAMS": "LAR", "LA": "LAR",
    "LA C": "LAC", "SD": "LAC", "CHARGERS": "LAC",
    "JAX": "JAC", "JAG": "JAC",
    "WSH": "WAS", "WFT": "WAS",
    "NYG": "NYG", "NYJ": "NYJ",
    "GBP": "GB", "KCC": "KC", "NEP": "NE", "NOS": "NO", "SFO": "SF",
    "TBB": "TB", "OAK": "LV", "LVR": "LV",
}


def normalise(raw):
    """-> (team | None, problem | None)"""
    if raw is None:
        return None, None
    code = str(raw).strip().upper()
    if not code:
        return None, None
    code = ALIASES.get(code, code)
    if code in TEAMS:
        return code, None
    return None, str(raw).strip()


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    src, year = sys.argv[1], int(sys.argv[2])

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sheet1"]

    entries, unknown = [], Counter()
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        entry_no = row[0].value
        nick = row[2].value
        if entry_no is None and not nick:
            continue

        picks, problems = {}, []
        for col in range(FIRST_WEEK_COL, MAX_WEEK_COL + 1):
            week = col - FIRST_WEEK_COL + 1
            team, bad = normalise(row[col - 1].value)
            if team:
                picks[week] = team
            elif bad:
                problems.append(f"w{week}={bad}")
                unknown[bad] += 1

        entries.append(
            {
                "entry": entry_no,
                "name": row[1].value,
                "nick": nick,
                "picks": picks,
                "problems": problems,
            }
        )

    played = sorted({w for e in entries for w in e["picks"]})
    if not played:
        sys.exit("no picks found — is this a blank template?")
    current = max(played)

    # Alive = submitted a pick in the most recent week with any activity.
    # A blank in an earlier week means eliminated, not skipped.
    for e in entries:
        e["alive"] = current in e["picks"]
        e["used"] = sorted(set(e["picks"].values()))

    weekly = {}
    for week in played:
        counts = Counter(e["picks"][week] for e in entries if week in e["picks"])
        total = sum(counts.values())
        weekly[week] = {
            "entrants": total,
            "distinctTeams": len(counts),
            "picks": [
                {"team": t, "count": n, "pct": round(100 * n / total, 1)}
                for t, n in counts.most_common()
            ],
        }

    out = ROOT / "data" / f"survivor-{year}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(
            {"year": year, "currentWeek": current, "weeks": weekly, "entries": entries},
            indent=1,
        ),
        encoding="utf-8",
    )

    for week in played:
        w = weekly[week]
        top = w["picks"][0]
        cum5 = sum(p["pct"] for p in w["picks"][:5])
        print(
            f"week {week:2d}: {w['entrants']:3d} alive, {w['distinctTeams']:2d} teams used | "
            f"top {top['team']} {top['pct']}% | top-5 {cum5:.1f}%"
        )

    print(f"\nwrote {out.relative_to(ROOT)}")
    if unknown:
        print("\nUNRECOGNISED TEAM CODES — add to ALIASES before trusting used-team lists:")
        for code, n in unknown.most_common():
            print(f"  {code!r} ×{n}")
        sys.exit(1)


if __name__ == "__main__":
    main()
