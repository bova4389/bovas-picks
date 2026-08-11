#!/usr/bin/env python3
"""
Parse Mike Lowe's "Weekly picks" workbook — the file he mails each Sunday once
games kick off, containing every entrant's card for the week.

This is the highest-value input we have. It gives the *actual* pick popularity of
the actual 270-person field, which is half of the leverage calculation in
STRATEGY.md §4. No national proxy comes close.

Sheet layout (one sheet per week, named "1".."18"):

    row 2, D..S      answer key — the winning NUMBER for each game, filled in
                     as results come in (blank in the Sunday-morning copy)
    row 2, T         "Pts"  (Monday-night tiebreaker column header)
    row 2, U..AJ     per-game headers "1/2", "3/4", ...
    row 3+, A        entry number
    row 3+, B        entrant real name
    row 3+, C        entrant nickname
    row 3+, D..S     that entrant's picked numbers, one per game
    row 3+, T        their Monday-night total-points guess
    row 3+, U..AJ    =IF(pick=answer,1,0) per game
    row 3+, AK       week total correct

Two outputs, split by shape rather than by sensitivity:

      data/raw/entries-<year>-w<NN>.json        names + individual cards, for
                                                the lookback "who picks well"
      data/popularity/pop-<year>-w<NN>.json     per-game aggregate percentages,
                                                for the leverage calculation

Entrant names are fine to commit here — see CLAUDE.md "Data & Privacy". The
source workbook itself is gitignored because its footer carries the
commissioner's email and cell number; nothing this script writes contains them.

    python scripts/parse_pool_picks.py "path/to/Weekly picks 25.xlsx" 2025 [week]
"""

import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PICK_COLS = range(4, 20)   # D..S — up to 16 games
NAME_COL, NICK_COL, MNF_COL, TOTAL_COL = 2, 3, 20, 37   # B, C, T, AK


def load_number_map(year):
    path = ROOT / "data" / f"number-map-{year}.json"
    if not path.exists():
        sys.exit(f"missing {path.relative_to(ROOT)} — run parse_weekly_sheets.py first")
    return json.loads(path.read_text(encoding="utf-8"))["weeks"]


def read_week(ws):
    """-> (answer_key, entries). Answer key is {} until results are posted."""
    key = {}
    for col in PICK_COLS:
        v = ws.cell(row=2, column=col).value
        if isinstance(v, int):
            key[col] = v

    entries = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if not isinstance(row[0].value, int):
            continue
        picks = [row[c - 1].value for c in PICK_COLS]
        picks = [p for p in picks if isinstance(p, int)]
        if not picks:
            continue  # blank template row
        entries.append(
            {
                "entry": row[0].value,
                "name": row[NAME_COL - 1].value,
                "nick": row[NICK_COL - 1].value,
                "picks": picks,
                "mnf": row[MNF_COL - 1].value,
                "correct": row[TOTAL_COL - 1].value if key else None,
            }
        )
    return key, entries


def popularity(games, entries):
    """Per-game share of the field on each side. Contains no personal data."""
    out = []
    for g in games:
        a, h = g["awayNum"], g["homeNum"]
        ca = sum(1 for e in entries if a in e["picks"])
        ch = sum(1 for e in entries if h in e["picks"])
        total = ca + ch
        if not total:
            continue
        out.append(
            {
                "awayNum": a, "away": g["away"],
                "homeNum": h, "home": g["home"],
                "awayCount": ca, "homeCount": ch,
                "awayPct": round(100 * ca / total, 1),
                "homePct": round(100 * ch / total, 1),
                # majority share — how much of the field agreed. High values mean
                # no leverage is available; the interesting games sit near 50.
                "concentration": round(100 * max(ca, ch) / total, 1),
                "majority": g["away"] if ca >= ch else g["home"],
            }
        )
    return out


def mnf_histogram(entries):
    """Tiebreaker guess density. Picking a local minimum wins outright ties."""
    vals = [e["mnf"] for e in entries if isinstance(e["mnf"], (int, float))]
    if not vals:
        return None
    counts = Counter(int(v) for v in vals)
    lo, hi = min(counts), max(counts)
    return {
        "n": len(vals),
        "min": lo, "max": hi,
        "histogram": {str(v): counts.get(v, 0) for v in range(lo, hi + 1)},
    }


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    src, year = sys.argv[1], int(sys.argv[2])
    only = int(sys.argv[3]) if len(sys.argv) > 3 else None

    number_map = load_number_map(year)
    wb = openpyxl.load_workbook(src, data_only=True)
    (ROOT / "data" / "raw").mkdir(parents=True, exist_ok=True)
    (ROOT / "data" / "popularity").mkdir(parents=True, exist_ok=True)

    for name in wb.sheetnames:
        if not name.isdigit():
            continue
        week = int(name)
        if only and week != only:
            continue
        if str(week) not in number_map:
            continue

        key, entries = read_week(wb[name])
        if not entries:
            continue  # blank template — weeks not yet played

        games = [g for g in number_map[str(week)] if g["counts"]]
        pop = popularity(games, entries)

        raw_path = ROOT / "data" / "raw" / f"entries-{year}-w{week:02d}.json"
        raw_path.write_text(
            json.dumps({"year": year, "week": week, "answerKey": key,
                        "entries": entries}, indent=1),
            encoding="utf-8",
        )

        pop_path = ROOT / "data" / "popularity" / f"pop-{year}-w{week:02d}.json"
        pop_path.write_text(
            json.dumps({"year": year, "week": week, "entrants": len(entries),
                        "games": pop, "mnf": mnf_histogram(entries)}, indent=1),
            encoding="utf-8",
        )

        conc = [g["concentration"] for g in pop]
        avg = sum(conc) / len(conc) if conc else 0
        print(
            f"week {week:2d}: {len(entries):3d} entries, {len(pop):2d} games, "
            f"mean concentration {avg:.1f}%, "
            f"{sum(1 for c in conc if c > 85)} games >85% chalk"
            + ("" if key else "   (no results yet)")
        )

    print("\nraw entries -> data/raw/          (names + cards, for lookback)")
    print("popularity  -> data/popularity/    (aggregates, for leverage calc)")


if __name__ == "__main__":
    main()
