#!/usr/bin/env python3
"""
Parse Mike Lowe's "Weekly Sheets" workbook into data/number-map-<year>.json.

He mails this once at the start of the season. It defines, for every week, which
games are in the pool and what number each team carries. Picks are submitted by
emailing those numbers, so this file is both the source of truth for the pick
form and the validator for outgoing picks.

Sheet layout (one sheet per week, named "1".."18"; any other sheet name is a
blank template and is skipped — 2025 mailed "30"/"32"/"Thanks", 2026 mailed
"24"/"26"/"28"/"30"/"32"):

    B1                     "Week 5"
    col A                  day header: Wednesday / Thursday / Friday /
                           Saturday / Sunday / Monday
    col B / C              away number ("13)") / away team
    col E / F              home number ("14)") / home team
    games with no number   excluded from the pool that week (note in col B)

Numbering is sequential in pairs down the sheet: away = odd, home = even.

The sheet never labels the Monday-night tiebreaker. See TIEBREAKERS below for
how it is derived, and for the one case that cannot be.

Output contains no personal information and is safe to commit.

    python scripts/parse_weekly_sheets.py "path/to/Weekly Sheets 25.xlsx" 2025
"""

import json
import sys
from pathlib import Path

import openpyxl

DAYS = {"Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "Monday"}
ROOT = Path(__file__).resolve().parent.parent

# The Monday-night tiebreaker — the game the "Monday Night Points" guess is
# scored against. The sheet does not label it, because in every week with a
# single Monday game there is nothing to label: the tiebreaker IS that game,
# and this parser derives it.
#
# A week with TWO Monday games is the case that cannot be derived. Only one of
# them carries the tiebreaker and the sheet does not say which, so guessing is
# worse than stopping — a wrong tiebreaker loses a tied week silently, and
# loses it in the one week the tiebreaker was needed. Such a week is a hard
# validation failure until its away number is recorded here.
#
#     TIEBREAKERS = {2026: {17: 25}}   # week 17 -> the game numbered 25/26
#
# Keyed by season so last year's answers cannot leak into this one.
TIEBREAKERS = {
    2026: {
        # Weeks 1-15 each have exactly one Monday game, so all fifteen are
        # derived and nothing needs recording. Weeks 16-18 are mailed later;
        # if either has a Monday doubleheader, put its away number here.
    },
}


def cell_values(row, n=6):
    """Row values as an A..F dict, tolerating merged cells."""
    vals = [c.value for c in row][:n]
    vals += [None] * (n - len(vals))
    return dict(zip("ABCDEF", vals))


def parse_num(v):
    """'13)' -> 13. Returns None if not a pick number."""
    if v is None:
        return None
    try:
        return int(str(v).strip().rstrip(")"))
    except ValueError:
        return None


def parse_week(ws):
    games, byes, day, in_byes = [], [], None, False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        c = cell_values(row)
        label = c["A"]
        if isinstance(label, str):
            head = label.strip().split()[0] if label.strip() else ""
            if head in DAYS:
                day = head   # "Thursday Nov" on holiday sheets

        away, home = c["C"], c["F"]

        # Sheets end with a "Bye Week" block listing bye teams in the same two
        # columns games use. Only real games carry "at" in column D, so that is
        # the discriminator — without it these were parsed as phantom games.
        if isinstance(away, str) and away.strip().lower().startswith("bye"):
            in_byes = True
            continue
        if not (isinstance(away, str) and isinstance(home, str)):
            continue
        if not away.strip() or not home.strip():
            continue
        if not (isinstance(c["D"], str) and "at" in c["D"].lower()):
            if in_byes:
                byes += [away.strip(), home.strip()]
            continue

        away_num, home_num = parse_num(c["B"]), parse_num(c["E"])
        games.append(
            {
                "day": day,
                "away": away.strip(),
                "home": home.strip(),
                # Unnumbered games are printed for reference but excluded from
                # scoring — normally the Thursday (and any Friday) game. On the
                # Thanksgiving sheet the Thursday games ARE numbered and count.
                "counts": away_num is not None and home_num is not None,
                "awayNum": away_num,
                "homeNum": home_num,
            }
        )
    return games, byes


def mark_tiebreaker(week_no, games, overrides):
    """Flag the one game the Monday-night points guess is scored against.

    Sets `tiebreaker` on every game so nothing downstream has to re-derive it,
    and so a hand-recorded answer is carried rather than guessed at again.
    Returns a problem string, or None.
    """
    for g in games:
        g["tiebreaker"] = False

    override = overrides.get(week_no)
    if override is not None:
        for g in games:
            if g["counts"] and g["awayNum"] == override:
                g["tiebreaker"] = True
                return None
        return (
            f"week {week_no}: TIEBREAKERS names away number {override}, "
            f"which is not a scored game on this sheet"
        )

    monday = [g for g in games if g["counts"] and g["day"] == "Monday"]
    if len(monday) == 1:
        monday[0]["tiebreaker"] = True
        return None
    if not monday:
        return f"week {week_no}: no scored Monday game — tiebreaker cannot be derived"

    listed = ", ".join(f"{g['away']} at {g['home']} ({g['awayNum']})" for g in monday)
    return (
        f"week {week_no}: {len(monday)} Monday games [{listed}] — only one is the "
        f"tiebreaker and the sheet does not say which. Add it to TIEBREAKERS "
        f"by away number."
    )


def validate(week_no, games):
    """Numbers must be a complete 1..2N sequence with away odd / home even."""
    problems = []
    scored = [g for g in games if g["counts"]]
    nums = sorted(n for g in scored for n in (g["awayNum"], g["homeNum"]))
    if nums != list(range(1, len(scored) * 2 + 1)):
        problems.append(f"week {week_no}: numbers are not a clean 1..{len(scored)*2} run: {nums}")
    for g in scored:
        if g["awayNum"] % 2 != 1 or g["homeNum"] != g["awayNum"] + 1:
            problems.append(
                f"week {week_no}: {g['away']} at {g['home']} numbered "
                f"{g['awayNum']}/{g['homeNum']} — expected odd away, home = away+1"
            )
    return problems


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    src, year = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True)
    overrides = TIEBREAKERS.get(int(year), {})
    weeks, byes, problems = {}, {}, []
    for name in wb.sheetnames:
        if not name.isdigit() or int(name) > 18:
            continue  # "24", "26", "28", "30", "32" are blank templates
        games, week_byes = parse_week(wb[name])
        if not any(g["counts"] for g in games):
            continue  # unfilled template sheet
        weeks[int(name)] = games
        byes[int(name)] = week_byes
        problems += validate(int(name), games)
        tb = mark_tiebreaker(int(name), games, overrides)
        if tb:
            problems.append(tb)

    out = ROOT / "data" / f"number-map-{year}.json"
    out.write_text(
        json.dumps({"year": int(year), "weeks": weeks, "byes": byes}, indent=1),
        encoding="utf-8",
    )

    for w in sorted(weeks):
        scored = [g for g in weeks[w] if g["counts"]]
        excl = [g for g in weeks[w] if not g["counts"]]
        top = max(g["homeNum"] for g in scored)
        note = f", excl {'/'.join(g['day'] or '?' for g in excl)}" if excl else ""
        bye = f", byes: {len(byes[w])}" if byes[w] else ""
        tb = next((g for g in scored if g["tiebreaker"]), None)
        tie = f" | TB {tb['away']} at {tb['home']} ({tb['awayNum']})" if tb else " | TB ??"
        print(f"week {w:2d}: {len(scored):2d} games, numbers 1-{top}{note}{bye}{tie}")

    print(f"\nwrote {out.relative_to(ROOT)}  ({len(weeks)} weeks)")
    if problems:
        print("\nVALIDATION PROBLEMS:")
        for p in problems:
            print("  !", p)
        sys.exit(1)
    print("validation: all weeks number cleanly")


if __name__ == "__main__":
    main()
